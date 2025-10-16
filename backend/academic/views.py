from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError
from django.http import HttpResponse
from django.db.models import Q
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import TextIOWrapper, BytesIO
from datetime import datetime

from .models import AnneeScolaire, Classe, Matiere, Eleve, MatiereClasse
from .serializers import (
    AnneeScolaireSerializer, ClasseSerializer, MatiereSerializer,
    EleveSerializer, EleveListSerializer, EleveImportSerializer,
    MatiereClasseSerializer
)
from users.models import Professeur
from users.permissions import IsAdminUser, IsTeacherOrAdmin, IsReadOnlyOrAdmin


class BaseEcoleViewSet(viewsets.ModelViewSet):
    """
    ViewSet de base avec filtrage automatique par école (Multi-Tenancy)
    Tous les ViewSets doivent hériter de cette classe
    """
    
    def get_queryset(self):
        """
        Filtre automatiquement les objets par l'école de l'utilisateur
        """
        queryset = super().get_queryset()
        
        # Si l'utilisateur est authentifié et a une école
        if self.request.user.is_authenticated:
            if hasattr(self.request, 'ecole') and self.request.ecole:
                # Filtrer par école (injectée par le middleware)
                return queryset.filter(ecole=self.request.ecole)
        
        # Si pas d'école, retourner vide (sécurité)
        return queryset.none()
    
    def perform_create(self, serializer):
        """
        Injecte automatiquement l'école lors de la création
        """
        if hasattr(self.request, 'ecole') and self.request.ecole:
            serializer.save(ecole=self.request.ecole)
        else:
            raise ValidationError({
                'error': 'Vous devez être assigné à une école pour créer cet objet',
                'code': 'NO_SCHOOL_ASSIGNED'
            })
    
    def perform_update(self, serializer):
        """
        Vérifie que l'objet appartient bien à l'école de l'utilisateur
        """
        instance = self.get_object()
        
        # Vérifier que l'objet appartient à l'école de l'utilisateur
        if hasattr(instance, 'ecole') and instance.ecole != self.request.ecole:
            raise ValidationError({
                'error': 'Vous ne pouvez pas modifier un objet d\'une autre école',
                'code': 'FORBIDDEN_CROSS_TENANT'
            })
        
        serializer.save()


class AnneeScolaireViewSet(BaseEcoleViewSet):
    """ViewSet pour la gestion des années scolaires - Admin uniquement pour modification"""
    queryset = AnneeScolaire.objects.all()
    serializer_class = AnneeScolaireSerializer
    permission_classes = [IsReadOnlyOrAdmin]  # Enseignants: lecture seule, Admin: tout
    
    # Le filtrage par école est automatique via BaseEcoleViewSet
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """Obtenir l'année scolaire active"""
        # Filtrer par école automatiquement
        annee = self.get_queryset().filter(active=True).first()
        if annee:
            serializer = self.get_serializer(annee)
            return Response(serializer.data)
        return Response(
            {'error': 'Aucune année scolaire active'},
            status=status.HTTP_404_NOT_FOUND
        )


class ClasseViewSet(BaseEcoleViewSet):
    """ViewSet pour la gestion des classes - Admin: tout, Enseignant: sa classe uniquement"""
    queryset = Classe.objects.all()
    serializer_class = ClasseSerializer
    permission_classes = [IsTeacherOrAdmin]
    
    def get_queryset(self):
        # Filtrage par école automatique via BaseEcoleViewSet
        queryset = super().get_queryset()
        
        user = self.request.user
        if user.is_admin():
            return queryset  # Admin voit toutes les classes de son école
        
        # Les professeurs ne voient que leurs classes
        try:
            professeur = user.professeur_profile
            return queryset.filter(
                Q(professeur_principal=professeur) |
                Q(matieres_enseignees__professeur=professeur)
            ).distinct()
        except:
            return queryset.none()
    
    @action(detail=True, methods=['get'])
    def eleves(self, request, pk=None):
        """Liste des élèves d'une classe"""
        classe = self.get_object()
        eleves = classe.eleves.filter(statut='actif')
        serializer = EleveListSerializer(eleves, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def add_matiere(self, request, pk=None):
        """Ajouter une matière à une classe"""
        if not request.user.is_admin():
            return Response(
                {'error': 'Permission refusée'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        classe = self.get_object()
        serializer = MatiereClasseSerializer(data={
            **request.data,
            'classe': classe.id
        })
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class MatiereViewSet(BaseEcoleViewSet):
    """ViewSet pour la gestion des matières - Admin: tout, Enseignant: lecture seule"""
    queryset = Matiere.objects.all()
    serializer_class = MatiereSerializer
    permission_classes = [IsReadOnlyOrAdmin]
    
    # Le filtrage par école est automatique via BaseEcoleViewSet


class EleveViewSet(BaseEcoleViewSet):
    """
    ViewSet pour la gestion des élèves
    - Admin: Peut tout faire (CRUD complet)
    - Enseignant: Peut seulement voir les élèves de SA classe (lecture seule)
    """
    queryset = Eleve.objects.all()
    permission_classes = [IsTeacherOrAdmin]
    
    def get_permissions(self):
        """
        Permissions dynamiques selon l'action :
        - Lecture (list, retrieve): Enseignants et Admins
        - Écriture (create, update, delete): Admin uniquement
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'import_csv']:
            permission_classes = [IsAdminUser]
        else:
            permission_classes = [IsTeacherOrAdmin]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrer par classe si spécifié
        classe_id = self.request.query_params.get('classe')
        if classe_id:
            queryset = queryset.filter(classe_id=classe_id)
        
        # Filtrer par statut
        statut = self.request.query_params.get('statut')
        if statut and statut != 'tous':
            queryset = queryset.filter(statut=statut)
        # Si statut='tous' ou non spécifié en mode passage, afficher tous
        # Sinon, par défaut, filtrer sur actifs seulement pour la page Élèves classique
        
        # Recherche
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nom__icontains=search) |
                Q(prenom__icontains=search) |
                Q(matricule__icontains=search)
            )
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return EleveListSerializer
        return EleveSerializer
    
    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Importer des élèves depuis un fichier CSV ou Excel"""
        if not request.user.is_admin():
            return Response(
                {'error': 'Seuls les administrateurs peuvent importer des élèves'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = EleveImportSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        file = serializer.validated_data['file']
        classe_id = serializer.validated_data['classe_id']
        
        try:
            classe = Classe.objects.get(id=classe_id)
        except Classe.DoesNotExist:
            return Response(
                {'error': 'Classe introuvable'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        imported = 0
        errors = []
        
        try:
            if file.name.endswith('.csv'):
                # Import CSV
                decoded_file = TextIOWrapper(file.file, encoding='utf-8')
                reader = csv.DictReader(decoded_file)
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        self._create_eleve_from_row(row, classe)
                        imported += 1
                    except Exception as e:
                        errors.append(f"Ligne {row_num}: {str(e)}")
            
            elif file.name.endswith(('.xlsx', '.xls')):
                # Import Excel
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_dict = dict(zip(headers, row))
                        self._create_eleve_from_row(row_dict, classe)
                        imported += 1
                    except Exception as e:
                        errors.append(f"Ligne {row_num}: {str(e)}")
            
            return Response({
                'success': True,
                'imported': imported,
                'errors': errors
            })
        
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de l\'import: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _create_eleve_from_row(self, row, classe):
        """Créer un élève à partir d'une ligne de données"""
        # Parse date de naissance
        date_naissance_str = row.get('date_naissance', '')
        if isinstance(date_naissance_str, str):
            date_naissance = datetime.strptime(date_naissance_str, '%Y-%m-%d').date()
        else:
            date_naissance = date_naissance_str
        
        # Générer matricule si non fourni
        matricule = row.get('matricule')
        if not matricule:
            last_eleve = Eleve.objects.order_by('-id').first()
            next_id = (last_eleve.id + 1) if last_eleve else 1
            matricule = f"EL{next_id:05d}"
        
        eleve = Eleve.objects.create(
            matricule=matricule,
            nom=row.get('nom', '').upper(),
            prenom=row.get('prenom', '').title(),
            sexe=row.get('sexe', 'M'),
            date_naissance=date_naissance,
            lieu_naissance=row.get('lieu_naissance', ''),
            telephone_eleve=row.get('telephone_eleve', ''),
            email=row.get('email', ''),
            adresse=row.get('adresse', ''),
            nom_pere=row.get('nom_pere', ''),
            telephone_pere=row.get('telephone_pere', ''),
            nom_mere=row.get('nom_mere', ''),
            telephone_mere=row.get('telephone_mere', ''),
            tuteur=row.get('tuteur', ''),
            telephone_tuteur=row.get('telephone_tuteur', ''),
            classe=classe,
            statut='actif'
        )
        return eleve
    
    @action(detail=False, methods=['get'])
    def template_csv(self, request):
        """Télécharger un modèle CSV pour l'import"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="template_import_eleves.csv"'
        
        # Ajouter BOM pour Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Ligne d'instructions
        writer.writerow(['# TEMPLATE IMPORT ÉLÈVES - Remplissez les lignes suivantes avec vos données'])
        writer.writerow(['# Format date: AAAA-MM-JJ | Sexe: M ou F | Les champs vides sont autorisés'])
        writer.writerow([])  # Ligne vide
        
        # En-têtes avec descriptions
        writer.writerow([
            'Matricule*', 
            'Nom*', 
            'Prénom*', 
            'Sexe* (M/F)', 
            'Date Naissance (AAAA-MM-JJ)', 
            'Lieu Naissance',
            'Téléphone Élève', 
            'Email', 
            'Adresse',
            'Nom Père', 
            'Téléphone Père', 
            'Nom Mère', 
            'Téléphone Mère',
            'Nom Tuteur', 
            'Téléphone Tuteur'
        ])
        
        # Exemples avec différents cas
        writer.writerow([
            'EL2024001', 'DIOP', 'Amadou', 'M', '2012-03-15', 'Dakar',
            '77 123 45 67', 'amadou.diop@email.com', 'Parcelles Assainies, Villa 123',
            'Moussa DIOP', '77 111 11 11', 'Fatou DIOP', '77 222 22 22',
            '', ''
        ])
        
        writer.writerow([
            'EL2024002', 'NDIAYE', 'Awa', 'F', '2011-08-22', 'Thiès',
            '78 234 56 78', '', 'Cité Keur Gorgui, Maison 45',
            'Ibrahima NDIAYE', '78 333 33 33', 'Aissatou NDIAYE', '78 444 44 44',
            '', ''
        ])
        
        writer.writerow([
            'EL2024003', 'FALL', 'Cheikh', 'M', '2013-01-10', 'Saint-Louis',
            '', '', 'Médina, Rue 15',
            'Omar FALL', '70 555 55 55', '', '',
            'Khadija SARR', '70 666 66 66'
        ])
        
        writer.writerow([])  # Ligne vide
        writer.writerow(['# (*) Champs obligatoires | Autres champs optionnels'])
        writer.writerow(['# Supprimez ces lignes de commentaires avant import'])
        
        return response
    
    @action(detail=False, methods=['get'])
    def template_excel(self, request):
        """Télécharger un modèle Excel pour l'import"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Élèves"
        
        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        instruction_font = Font(italic=True, size=10)
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Instructions
        ws.merge_cells('A1:O1')
        ws['A1'] = "📋 TEMPLATE IMPORT ÉLÈVES - Remplissez les lignes 5 et suivantes avec vos données"
        ws['A1'].fill = instruction_fill
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:O2')
        ws['A2'] = "Format date: AAAA-MM-JJ | Sexe: M ou F | (*) = Champs obligatoires"
        ws['A2'].fill = instruction_fill
        ws['A2'].font = instruction_font
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # En-têtes (ligne 4)
        headers = [
            'Matricule*', 'Nom*', 'Prénom*', 'Sexe*', 'Date Naissance', 
            'Lieu Naissance', 'Téléphone Élève', 'Email', 'Adresse',
            'Nom Père', 'Téléphone Père', 'Nom Mère', 'Téléphone Mère',
            'Nom Tuteur', 'Téléphone Tuteur'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Exemples
        examples = [
            ['EL2024001', 'DIOP', 'Amadou', 'M', '2012-03-15', 'Dakar',
             '77 123 45 67', 'amadou.diop@email.com', 'Parcelles Assainies, Villa 123',
             'Moussa DIOP', '77 111 11 11', 'Fatou DIOP', '77 222 22 22', '', ''],
            
            ['EL2024002', 'NDIAYE', 'Awa', 'F', '2011-08-22', 'Thiès',
             '78 234 56 78', '', 'Cité Keur Gorgui, Maison 45',
             'Ibrahima NDIAYE', '78 333 33 33', 'Aissatou NDIAYE', '78 444 44 44', '', ''],
            
            ['EL2024003', 'FALL', 'Cheikh', 'M', '2013-01-10', 'Saint-Louis',
             '', '', 'Médina, Rue 15',
             'Omar FALL', '70 555 55 55', '', '', 'Khadija SARR', '70 666 66 66']
        ]
        
        for row_num, example in enumerate(examples, 5):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = example_fill
        
        # Ajuster largeur colonnes
        column_widths = [12, 15, 15, 8, 15, 15, 15, 25, 30, 15, 15, 15, 15, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Fixer hauteur lignes
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[4].height = 35
        
        # Note en bas
        ws.merge_cells('A10:O10')
        ws['A10'] = "💡 Supprimez les exemples (lignes 5-7) et ajoutez vos propres élèves à partir de la ligne 5"
        ws['A10'].fill = instruction_fill
        ws['A10'].font = instruction_font
        ws['A10'].alignment = Alignment(horizontal='center')
        
        # Sauvegarder
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_import_eleves.xlsx"'
        
        return response
    
    @action(detail=True, methods=['patch'])
    def proposer_passage(self, request, pk=None):
        """
        Enseignant peut proposer le statut (admis/redouble) pour un élève
        Accessible aux profs principaux
        """
        eleve = self.get_object()
        nouveau_statut = request.data.get('statut')
        
        # Vérifier que c'est le prof principal de la classe de l'élève
        if request.user.is_professeur() and not request.user.is_admin():
            prof = Professeur.objects.get(user=request.user)
            if eleve.classe.professeur_principal != prof:
                return Response(
                    {'error': 'Vous n\'êtes pas le professeur principal de cet élève'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        if nouveau_statut not in ['admis', 'redouble', 'actif']:
            return Response(
                {'error': 'Statut invalide. Utilisez: admis, redouble ou actif'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        eleve.statut = nouveau_statut
        eleve.save()
        
        return Response({
            'success': True,
            'message': f'Statut de {eleve.nom_complet} changé en "{nouveau_statut}"',
            'eleve': EleveListSerializer(eleve).data
        })
    
    @action(detail=False, methods=['post'])
    def passage_classe(self, request):
        """
        ADMIN UNIQUEMENT: Effectuer le passage de classe officiel
        Change la classe ET le statut des élèves
        Payload: {
            "eleves": [id1, id2, ...],
            "nouvelle_classe": classe_id,
            "statut": "actif"
        }
        """
        if not request.user.is_admin():
            return Response(
                {'error': 'Action réservée aux administrateurs (passage de classe officiel)'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        eleves_ids = request.data.get('eleves', [])
        nouvelle_classe_id = request.data.get('nouvelle_classe')
        nouveau_statut = request.data.get('statut', 'actif')
        
        if not eleves_ids or not nouvelle_classe_id:
            return Response(
                {'error': 'IDs des élèves et nouvelle classe requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            nouvelle_classe = Classe.objects.get(id=nouvelle_classe_id)
            eleves = Eleve.objects.filter(id__in=eleves_ids)
            
            # Mise à jour
            eleves_mis_a_jour = 0
            for eleve in eleves:
                eleve.classe = nouvelle_classe
                eleve.statut = nouveau_statut
                eleve.save()
                eleves_mis_a_jour += 1
            
            return Response({
                'success': True,
                'message': f'{eleves_mis_a_jour} élève(s) déplacé(s) vers {nouvelle_classe.nom}',
                'eleves_mis_a_jour': eleves_mis_a_jour
            })
            
        except Classe.DoesNotExist:
            return Response(
                {'error': 'Classe introuvable'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MatiereClasseViewSet(BaseEcoleViewSet):
    """ViewSet pour la gestion des matières par classe"""
    queryset = MatiereClasse.objects.all()
    serializer_class = MatiereClasseSerializer
    permission_classes = [IsTeacherOrAdmin]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrer par classe
        classe_id = self.request.query_params.get('classe')
        if classe_id:
            queryset = queryset.filter(classe_id=classe_id)
        
        return queryset
